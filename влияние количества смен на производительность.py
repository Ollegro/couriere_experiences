import matplotlib
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
matplotlib.use('Agg')  # Отключаем интерактивный режим matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from openpyxl.styles import Font

# Загрузка данных
df_reset = pd.read_csv('results/courier_score.csv').reset_index()

# Определение бинов и меток
bins = [20, 50, 100, 200, 500]
labels = ['20-50', '51-100', '101-200', '201-500']

# Добавляем столбец с категорией смен
df_reset['shift_group'] = pd.cut(df_reset['total_shifts'], bins=bins, labels=labels)

# Группируем данные
grouped = df_reset.groupby('shift_group', observed=True).agg(
    avg_orders=('avg_orders_per_shift', 'mean'),
    courier_count=('id_driver', 'count')
).round(2)

# Создаем буфер для графика
buf = BytesIO()

# Строим график
fig, ax = plt.subplots(figsize=(8, 5))
grouped['avg_orders'].plot(kind='bar', ax=ax, color='skyblue')
plt.title('Среднее количество заказов в смену в зависимости от общего числа смен')
plt.xlabel('Число смен (total_shifts)')
plt.ylabel('Среднее число заказов в смену')
plt.xticks(rotation=45)

# Добавляем подписи данных
for i, (index, row) in enumerate(grouped.iterrows()):
    ax.text(i, row['avg_orders'] + 0.1, f'{row["avg_orders"]:.2f}',
            ha='center', fontsize=11, fontweight='bold')
    ax.text(i, row['avg_orders'] - 0.15, f'n={row["courier_count"]}',
            ha='center', fontsize=9, color='dimgray')

plt.figtext(0.5, 0.01, 'n = количество курьеров в группе',
            ha='center', fontsize=10, color='gray', style='italic')
plt.ylim(grouped['avg_orders'].min() - 0.3, grouped['avg_orders'].max() + 0.2)
plt.grid(True, axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()

# Сохраняем график в буфер
plt.savefig(buf, format='png', dpi=150)
buf.seek(0)
plt.close(fig)

# Сохранение в Excel
output_path = 'results/courier_analysis.xlsx'
sheet_name = 'заказы_в_смену'

try:
    wb = load_workbook(output_path)
except FileNotFoundError:
    wb = Workbook()

# Удаляем старый лист, если он существует
if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])

ws = wb.create_sheet(sheet_name)

# Записываем заголовок
ws['A1'] = "Влияние количества смен на объем исполненных заказов в смену"
ws['A1'].font = Font(bold=True, size=12)

# Записываем таблицу
for r in dataframe_to_rows(grouped.reset_index(), index=False, header=True):
    ws.append(r)

# Добавляем график
img = XLImage(buf)
ws.add_image(img, 'A10')  # Размещаем график ниже таблицы

# Автоподбор ширины столбцов
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
wb.save(output_path)

print("Анализ успешно сохранен в файл courier_analysis.xlsx на лист 'заказы_в_смену'")