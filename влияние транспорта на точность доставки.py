import matplotlib
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

# Установка бэкенда Agg для работы без GUI
matplotlib.use('Agg')

# Загрузка данных
df = pd.read_csv('results/courier_score.csv').reset_index()

# Создаем буфер для графика
buf = BytesIO()

# Настройки графика
plt.figure(figsize=(10, 6))

# Уникальные типы транспорта
transport_types = df['courier_dtype'].unique()

# Для каждого типа транспорта строим гистограмму
for transport in transport_types:
    subset = df[df['courier_dtype'] == transport]['pct_timer_up_expired']
    # Веса: каждый элемент — доля в процентах от общего числа курьеров
    weights = np.ones_like(subset) / len(subset) * 100
    # Строим гистограмму как контур (без заливки)
    plt.hist(subset, bins=20,
             weights=weights,
             histtype='step',  # только контур
             linewidth=2,  # толщина линии
             label=transport)

# Оформление графика
plt.title('Процент курьеров по уровню просроченных заказов')
plt.xlabel('Процент просроченных заказов')
plt.ylabel('Процент курьеров (%)')
plt.legend(title='Тип транспорта')
plt.grid(True, linestyle='--', alpha=0.5)
plt.tight_layout()

# Сохраняем график в буфер
plt.savefig(buf, format='png', dpi=150)
buf.seek(0)
plt.close()

# Подготовка данных для таблицы
stats_data = []
for transport in transport_types:
    subset = df[df['courier_dtype'] == transport]['pct_timer_up_expired']
    stats_data.append({
        'Тип транспорта': transport,
        'Количество курьеров': len(subset),
        'Средний % просрочек': subset.mean(),
        'Медианный % просрочек': subset.median()
    })

stats_df = pd.DataFrame(stats_data).round(2)

# Сохранение в Excel
output_path = 'results/courier_analysis.xlsx'
sheet_name = 'просрочки_по_типам'

try:
    wb = load_workbook(output_path)
except FileNotFoundError:
    wb = Workbook()

# Удаляем старый лист, если он существует
if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])

ws = wb.create_sheet(sheet_name)

# Записываем заголовок
ws['A1'] = "Анализ просроченных заказов по типам транспорта"
ws['A1'].font = Font(bold=True, size=12)

# Записываем таблицу со статистикой
for r in dataframe_to_rows(stats_df, index=False, header=True):
    ws.append(r)

# Добавляем график
img = XLImage(buf)
ws.add_image(img, 'A10')  # Размещаем график ниже таблицы

# Автоподбор ширины столбцов
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
wb.save(output_path)

print("Анализ успешно сохранен в файл courier_analysis.xlsx на лист 'просрочки_по_типам'")