import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.reader.excel import load_workbook
from sklearn.preprocessing import MinMaxScaler
from data import df
matplotlib.use('TkAgg')
pd.set_option('display.width', None)  # Автоматически подбирает ширину под консоль
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import os

# Основные факторы для расчёта КПД:
# avg_orders_per_shift  Среднее число заказов за смену  ↑ больше = ↑ эффективность
# days_active  Дней активности  ↑ больше = ↑ надежность
# total_shifts  Общее количество смен  ↑ больше = ↑ опытность
# median_time_ratio - медиана факта времени к плану. <1 - работает быстрее планового, >1 - наоборот
# avg_time_ratio - среднее факта времени к плану. Не учитывает выбросы
# pct_timer_up_expired  Процент просроченных доставок  ↓ меньше = ↑ качество
# stddev_up_ratio , iqr_up_ratio  Мера нестабильности доставок  ↓ меньше = ↑ предсказуемость Не используем !!!!!!!

# Метрики для оценки КПД
metrics = [
    'avg_orders_per_shift',
    'median_time_ratio',
    'pct_timer_up_expired',
    'stddev_up_ratio',
    'days_active'
]

# Нормализация
scaler = MinMaxScaler()
df_scaled = pd.DataFrame(scaler.fit_transform(df[metrics]), columns=metrics, index=df.index)

# Веса
weights = {
    'avg_orders_per_shift': 0.5,
    'median_time_ratio': -0.1,  # обратная метрика
    'pct_timer_up_expired': -0.1,
    'stddev_up_ratio': -0.1,
    'days_active': 0.1  # стабильность
}

# Расчёт courier_score
df['courier_score'] = (
        df_scaled['avg_orders_per_shift'] * weights['avg_orders_per_shift'] +
        (1 - df_scaled['median_time_ratio']) * weights['median_time_ratio'] +
        (1 - df_scaled['pct_timer_up_expired']) * weights['pct_timer_up_expired'] +
        (1 - df_scaled['stddev_up_ratio']) * weights['stddev_up_ratio'] +
        df_scaled['days_active'] * weights['days_active']
)

# Нормализуем courier_score от 0 до 1
df['courier_score'] = MinMaxScaler().fit_transform(df[['courier_score']])

# Переставляем колонки так, чтобы courier_score был первой колонкой после индекса
cols = ['courier_score'] + [col for col in df.columns if col != 'courier_score']
df = df[cols]

# Сортируем и выводим
print('Датафрейм с посчитанным КПД - courier_score','\n',df.sort_values('courier_score', ascending=False), '\n')

# Записываем датафрейм в новый csv
df.to_csv('results/courier_score.csv')


df = pd.read_csv('results/courier_score.csv')
# # Топ N самых эффективных курьеров
top_n = 10
top_couriers = df['courier_score'].sort_values(ascending=False).head(top_n)

# ГРУППИРОВКА ПО КАТЕГОРИИ ОПЫТА

# Убедимся что индекс не мешает
df_reset = df.reset_index()

# Группируем по experience_category и считаем средний courier_score
grouped_by_exp = df_reset.groupby('experience_category', observed=True).agg(
    avg_courier_score=('courier_score', 'mean'),
    courier_count=('id_driver', 'count')
).sort_values(by='avg_courier_score', ascending=False).round(2)
# print('Зависимость КПД от опыта', '\n',grouped_by_exp, '\n')


# вычислить датафрейм и нарисовать гистограмму зависимость КПД от количества смен
# Восстанавливаем id_driver как столбец для работы
df_reset = df.reset_index()

# Создаём бины и метки для категоризации
bins = [20, 49, 99, 199, float('inf')]
labels = ['20-49', '50-99', '100-199', '200+']

# Добавляем категорию "Количество смен"
df_reset['shifts_category'] = pd.cut(df_reset['total_shifts'], bins=bins, labels=labels)

# Группируем по количеству смен и считаем средний courier_score и количество курьеров
grouped_by_shifts = df_reset.groupby('shifts_category', observed=True).agg(
    avg_courier_score=('courier_score', 'mean'),
    courier_count=('id_driver', 'count')
).round(2)

# Сохраняем все графики во временные файлы
plt.figure(figsize=(10, 6))
top_couriers.plot(kind='bar', color='skyblue')
plt.title('Топ 10 курьеров по КПД')
plt.xlabel('ID курьера')
plt.ylabel('КПД (courier_score)')
plt.xticks(rotation=45)
plt.grid(True)
plt.tight_layout()
plt.savefig('top_couriers.png')
plt.close()

plt.figure(figsize=(8, 5))
plt.hist(df['courier_score'], bins=30, color='teal', edgecolor='black')
plt.title('Распределение КПД курьеров')
plt.xlabel('courier_score')
plt.ylabel('Частота')
plt.grid(True)
plt.tight_layout()
plt.savefig('score_distribution.png')
plt.close()

plt.figure(figsize=(8, 6))
plt.scatter(df['avg_orders_per_shift'], df['courier_score'], alpha=0.6, c='green')
plt.title('Зависимость КПД от заказов в смену')
plt.xlabel('Среднее число заказов в смену')
plt.ylabel('courier_score')
plt.grid(True)
plt.tight_layout()
plt.savefig('orders_vs_score.png')
plt.close()

plt.figure(figsize=(8, 6))
plt.plot(grouped_by_exp.index, grouped_by_exp['avg_courier_score'], marker='o', color='darkcyan')
plt.title('Зависимость КПД от категории опыта')
plt.xlabel('Категория опыта')
plt.ylabel('avg_courier_score')
plt.xticks(rotation=45)
plt.grid(True)
plt.tight_layout()
plt.savefig('experience_vs_score.png')
plt.close()

# Создаем или загружаем файл Excel
excel_file = 'results/courier_analysis.xlsx'
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    if 'kpd_research' in wb.sheetnames:
        del wb['kpd_research']
else:
    wb = Workbook()
    del wb['Sheet']  # Удаляем дефолтный лист

ws = wb.create_sheet("kpd_research")

# 1. Основные данные
ws['A1'] = "Основные данные по КПД курьеров"
ws['A1'].font = Font(bold=True, size=12)
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 2. Группировка по опыту
ws.append([])
ws.append(["Группировка по категории опыта"])
for r in dataframe_to_rows(grouped_by_exp, index=True, header=True):
    ws.append(r)

# 3. Группировка по сменам
for _ in range(3):
    ws.append([])
ws.append(["Группировка по количеству смен"])
for r in dataframe_to_rows(grouped_by_shifts, index=True, header=True):
    ws.append(r)

# Вставляем графики
img_top = Image('top_couriers.png')
ws.add_image(img_top, 'A20')

img_dist = Image('score_distribution.png')
ws.add_image(img_dist, 'A60')

img_orders = Image('orders_vs_score.png')
ws.add_image(img_orders, 'A100')

img_exp = Image('experience_vs_score.png')
ws.add_image(img_exp, 'A140')

# Автонастройка ширины столбцов
for column in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column if cell.value)
    ws.column_dimensions[column[0].column_letter].width = max_length + 2

# Сохраняем и удаляем временные файлы
wb.save(excel_file)
for img_file in ['top_couriers.png', 'score_distribution.png',
                 'orders_vs_score.png', 'experience_vs_score.png']:
    os.remove(img_file)

print(f"Все данные и графики сохранены на лист 'kpd_research' в файле {excel_file}")