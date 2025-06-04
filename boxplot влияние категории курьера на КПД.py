import matplotlib
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

matplotlib.use('TkAgg')
pd.set_option('display.width', None)  # Автоматически подбирает ширину под консоль



# СТРОИМ ЯШИК С УСАМИ ПО КАТЕГОРИЯМ ОПЫТА ДЛЯ ЗНАЧЕНИЯ COURIER_SCORE
# Группируем данные по experience_category, указываем observed=True

df = pd.read_csv('results/courier_score.csv')


groups = df.groupby('experience_category', observed=True)['courier_score'].apply(list)

# Подготавливаем данные для boxplot
data = [np.array(groups[cat]) for cat in groups.index]

# Строим boxplot
plt.figure(figsize=(10, 6))
plt.boxplot(data, tick_labels=groups.index, patch_artist=True)  # <-- labels → tick_labels

# Оформление графика
colors = ['lightblue', 'lightgreen', 'salmon']
bplots = plt.boxplot(data, tick_labels=groups.index, patch_artist=True)

for patch, color in zip(bplots['boxes'], colors):
    patch.set_facecolor(color)

plt.title('Распределение courier_score по категориям опыта', fontsize=14)
plt.xlabel('Категория опыта', fontsize=12)
plt.ylabel('courier_score', fontsize=12)
plt.grid(True, linestyle='--', alpha=0.6)
plt.xticks(rotation=0)
plt.tight_layout()

# Показываем график
plt.show()

# ***********************************************************************************************************************************
# запись в эксель
import matplotlib
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# Настройки отображения
matplotlib.use('Agg')  # Для сохранения графиков без отображения
pd.set_option('display.width', None)

# Загрузка данных
df = pd.read_csv('results/courier_score.csv')

# 1. Подготовка данных для анализа
groups = df.groupby('experience_category', observed=True)['courier_score'].apply(list)
data = [np.array(groups[cat]) for cat in groups.index]

# 2. Создание boxplot
plt.figure(figsize=(6, 3.5))  # было (12, 7)
bplots = plt.boxplot(data, patch_artist=True)

# Настройка внешнего вида графика
colors = ['lightblue', 'lightgreen', 'salmon']
for patch, color in zip(bplots['boxes'], colors):
    patch.set_facecolor(color)

plt.title('Распределение courier_score по категориям опыта', fontsize=14)
plt.xlabel('Категория опыта', fontsize=12)
plt.ylabel('courier_score', fontsize=12)
plt.xticks(range(1, len(groups.index)+1), groups.index)
plt.grid(True, linestyle='--', alpha=0.6)
plt.tight_layout()

# Сохранение графика во временный файл
graph_path = 'temp_boxplot.png'
plt.savefig(graph_path, dpi=150, bbox_inches='tight')  # уменьшаем dpi с 300 до 150
plt.close()

# 3. Подготовка табличных данных
stats_df = df.groupby('experience_category', observed=True)['courier_score'].describe()
stats_df = stats_df[['count', 'mean', '50%', 'std', 'min', 'max']]
stats_df.columns = ['Количество', 'Среднее', 'Медиана', 'Станд. отклонение', 'Минимум', 'Максимум']

# 4. Сохранение в Excel
output_file = 'results/courier_analysis.xlsx'

# Создаем папку results, если её нет
os.makedirs('results', exist_ok=True)

# Загружаем или создаем файл Excel
try:
    book = load_workbook(output_file)
except FileNotFoundError:
    book = Workbook()
    book.remove(book.active)  # Удаляем дефолтный лист

# Удаляем старый лист, если существует
if "влияние категории на кпд" in book.sheetnames:
    book.remove(book["влияние категории на кпд"])

# Создаем новый лист
sheet = book.create_sheet("влияние категории на кпд")

# Записываем заголовок
sheet['A1'] = "Анализ влияния категории опыта на КПД курьеров"
sheet['A1'].font = Font(bold=True, size=14)

# Записываем таблицу с статистикой
sheet['A3'] = "Статистика по категориям опыта"
for r in dataframe_to_rows(stats_df.reset_index(), index=False, header=True):
    sheet.append(r)

# Вставляем график
img = Image(graph_path)
sheet.add_image(img, 'A10')  # Позиция графика

# Настраиваем ширину столбцов
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    sheet.column_dimensions[column].width = max_length + 2

# Сохраняем файл
book.save(output_file)

# Удаляем временный файл графика
os.remove(graph_path)

print(f"Результаты сохранены в файл {output_file} на лист 'влияние категории на кпд'")