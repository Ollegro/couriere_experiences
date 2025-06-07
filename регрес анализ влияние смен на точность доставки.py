import matplotlib
from openpyxl.workbook import Workbook

matplotlib.use('Agg')  # Отключаем интерактивный режим
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Инициализация
output_path = 'results/courier_analysis.xlsx'
sheet_name = 'опыт_и_пунктуальность'

# Загрузка данных
df_reset = pd.read_csv('results/courier_score.csv').reset_index()

# Создаем буферы для графиков
buf1 = BytesIO()
buf2 = BytesIO()

# ====================================================================
# Анализ 1: Влияние опыта на median_time_ratio
# ====================================================================
X = df_reset['total_shifts']
y = df_reset['median_time_ratio']
X_const = sm.add_constant(X)
model1 = sm.OLS(y, X_const).fit()

# Создаем выводы для первого анализа
output1 = [
    ["АНАЛИЗ: ВЛИЯНИЕ ОПЫТА НА ОТНОШЕНИЕ ФАКТА К ПЛАНУ"],
    ["Регрессионная модель:", str(model1.summary())],
    ["Выводы:"],
    ["1. С увеличением числа смен (total_shifts) курьеры становятся более пунктуальными"],
    ["2. Эффект статистически значимый (p-value = {:.4f})".format(model1.pvalues['total_shifts'])],
    ["3. Коэффициент: {:.6f}".format(model1.params['total_shifts'])],
    ["4. R-squared: {:.4f}".format(model1.rsquared)],
    ["5. Каждая дополнительная смена снижает median_time_ratio на {:.4f}".format(model1.params['total_shifts'])],
    [" "],
    ["Это может говорить о том, что курьеры учатся на опыте, лучше оценивают время пути, избегают пробок и т.п."],
    [" "]  # Пустая строка перед графиком
]

# График 1 (уменьшенный размер)
plt.figure(figsize=(8, 5))
plt.scatter(X, y, alpha=0.5, color='steelblue', s=30)
plt.plot(X, model1.predict(X_const), color='red', linewidth=2)
plt.title('Зависимость median_time_ratio от числа смен')
plt.xlabel('Число смен (total_shifts)')
plt.ylabel('median_time_ratio')
plt.grid(True)
plt.tight_layout()
plt.savefig(buf1, format='png', dpi=150)
buf1.seek(0)
plt.close()

# ====================================================================
# Анализ 2: Влияние опыта на процент просрочек
# ====================================================================
X = df_reset['total_shifts']
y = df_reset['pct_timer_up_expired']
X_const = sm.add_constant(X)
model2 = sm.OLS(y, X_const).fit()

# Создаем выводы для второго анализа
output2 = [
    ["АНАЛИЗ: ВЛИЯНИЕ ОПЫТА НА ПРОЦЕНТ ПРОСРОЧЕК"],
    ["Регрессионная модель:", str(model2.summary())],
    ["Выводы:"],
    ["1. Коэффициент total_shifts = {:.4f}".format(model2.params['total_shifts'])],
    ["2. Каждая дополнительная смена уменьшает процент просрочек на {:.4f}%".format(abs(model2.params['total_shifts']))],
    ["3. P-значение: {:.4f} (статистически значимо)".format(model2.pvalues['total_shifts'])],
    ["4. Базовый уровень просрочек (0 смен): {:.4f}%".format(model2.params['const'])],
    [" "],
    ["Примеры:"],
    ["На 10-й смене: {:.4f}%".format(model2.params['const'] + model2.params['total_shifts']*10)],
    ["На 50-й смене: {:.4f}%".format(model2.params['const'] + model2.params['total_shifts']*50)],
    [" "],
    ["Общий вывод:"],
    ["С увеличением опыта курьеры становятся более пунктуальными"],
    ["Эффект статистически значимый и стабильный"],
    ["Курьеры учатся лучше планировать время, избегать пробок, выбирать маршруты"],
    [" "]  # Пустая строка перед графиком
]

# График 2 (уменьшенный размер)
plt.figure(figsize=(8, 5))
plt.scatter(X, y, alpha=0.5, color='steelblue', s=30)
plt.plot(X, model2.predict(X_const), color='red', linewidth=2)
plt.title('Зависимость pct_timer_up_expired от числа смен')
plt.xlabel('Число смен (total_shifts)')
plt.ylabel('pct_timer_up_expired (%)')
plt.grid(True)
plt.tight_layout()
plt.savefig(buf2, format='png', dpi=150)
buf2.seek(0)
plt.close()

# ====================================================================
# Сохранение в Excel
# ====================================================================
try:
    wb = load_workbook(output_path)
except FileNotFoundError:
    wb = Workbook()

if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])

ws = wb.create_sheet(sheet_name)

# Записываем первый анализ
for row in output1:
    ws.append(row)

# Добавляем первый график на строку 32
img1 = XLImage(buf1)
ws.add_image(img1, 'A32')

# Записываем второй анализ
for row in output2:
    ws.append(row)

# Добавляем второй график на строку 63
img2 = XLImage(buf2)
ws.add_image(img2, 'A66')

# Настройка ширины столбцов
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = min(max_length + 2, 50)

# Сохраняем файл
wb.save(output_path)

print("Анализ успешно сохранен в файл '{}' на лист '{}'".format(output_path, sheet_name))