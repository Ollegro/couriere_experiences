import matplotlib
matplotlib.use('Agg')  # Отключаем интерактивный режим
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Загрузка данных
df_reset = pd.read_csv('results/courier_score.csv').reset_index()

# Группировка по courier_dtype
grouped = df_reset.groupby('courier_dtype').agg(
    avg_courier_score=('courier_score', 'mean'),
    courier_count=('id_driver', 'count'),
    avg_orders_per_shift=('avg_orders_per_shift', 'mean'),
    avg_median_time_ratio=('median_time_ratio', 'mean'),
    avg_pct_timer_up_expired=('pct_timer_up_expired', 'mean'),
    avg_stddev_up_ratio=('stddev_up_ratio', 'mean'),
    avg_days_active=('days_active', 'mean')
).round(2)

# Сортируем по КПД
grouped_sorted = grouped.sort_values('avg_courier_score')

# Создаем буфер для графика
buf = BytesIO()

# Строим график
plt.figure(figsize=(10, 6))
grouped_sorted['avg_courier_score'].plot(kind='bar', color='skyblue')
plt.title('Средний КПД курьеров по типу доставки')
plt.xlabel('Тип курьера (courier_dtype)')
plt.ylabel('avg_courier_score')
plt.xticks(rotation=45)
plt.grid(True)
plt.tight_layout()

# Сохраняем график в буфер
plt.savefig(buf, format='png', dpi=150)
buf.seek(0)
plt.close()

# Сохранение в Excel
output_path = 'results/courier_analysis.xlsx'
sheet_name = 'кпд_по_типам'

try:
    wb = load_workbook(output_path)
except FileNotFoundError:
    from openpyxl import Workbook
    wb = Workbook()

# Удаляем старый лист, если он существует
if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])

ws = wb.create_sheet(sheet_name)

# Записываем заголовок
ws['A1'] = "Сравнение КПД по типам курьеров"
ws['A1'].font = Font(bold=True, size=12)

# Записываем таблицу (начинаем с 3 строки)
for r in dataframe_to_rows(grouped_sorted.reset_index(), index=False, header=True):
    ws.append(r)

# Добавляем график
img = XLImage(buf)
ws.add_image(img, 'A' + str(len(grouped_sorted) + 5))  # Размещаем график ниже таблицы

# Настраиваем ширину столбцов
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
wb.save(output_path)

print("Анализ КПД по типам курьеров успешно сохранен в файл courier_analysis.xlsx на лист 'кпд_по_типам'")

