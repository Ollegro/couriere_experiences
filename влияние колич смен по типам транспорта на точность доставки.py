import matplotlib
matplotlib.use('Agg')  # Чтобы не рисовать график на экране
from openpyxl.styles import Font
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import statsmodels.api as sm
from statsmodels.tools.tools import add_constant
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows

# Загрузка данных
df = pd.read_csv('results/courier_score.csv')
df_reset = df.reset_index()
courier_types = df_reset['courier_dtype'].unique()

# Путь к файлу Excel
output_path = 'results/courier_analysis.xlsx'
sheet_name = 'смены_тип трансп на проср'

# Подготовка буфера для хранения графиков
graph_buffers = []

# Создание графиков и сбор результатов регрессии
results = []

plt.figure(figsize=(12, 8))

for i, dtype in enumerate(courier_types):
    df_group = df_reset[df_reset['courier_dtype'] == dtype]
    if len(df_group) < 2:
        continue

    X = df_group[['total_shifts']]
    y = df_group['pct_timer_up_expired']

    X_const = add_constant(X)
    model = sm.OLS(y, X_const).fit()

    results.append({
        'courier_dtype': dtype,
        'coefficient': model.params['total_shifts'],
        'p_value': model.pvalues['total_shifts'],
        'r_squared': model.rsquared,
        'sample_size': len(df_group)
    })

    # Рисуем график
    fig, ax = plt.subplots(figsize=(6, 4))
    sns.regplot(x='total_shifts', y='pct_timer_up_expired', data=df_group,
                scatter_kws={'alpha': 0.6}, line_kws={'color': 'red'}, ax=ax)

    ax.set_title(f'{dtype}\n(slope={model.params["total_shifts"]:.6f})')
    ax.set_xlabel('Число смен')
    ax.set_ylabel('pct_timer_up_expired')
    ax.grid(True)
    plt.tight_layout()

    # Сохраняем график в буфер
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150)
    buf.seek(0)
    graph_buffers.append((dtype, buf))
    plt.close(fig)

# Преобразуем в DataFrame
results_df = pd.DataFrame(results)

# **********************************************************************************************************************
# Сохранение в Excel с графиками и таблицей
from openpyxl import Workbook
from openpyxl.utils import dataframe as df2
from openpyxl.utils import get_column_letter

# Создаем или загружаем Excel-файл
try:
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
except FileNotFoundError:
    wb = Workbook()

# Удаляем старый лист, если он существует
if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])

# Создаем новый лист
ws = wb.create_sheet(sheet_name)

# Добавляем заголовок
ws['A1'] = "Регрессионный анализ: pct_timer_up_expired ~ total_shifts"
ws['A1'].font = Font(bold=True, size=12)

# Добавляем таблицу
for r in df2.dataframe_to_rows(results_df, index=False, header=True):
    ws.append(r)

# Добавляем выводы по каждому типу курьеров
ws.append(["Автомобильные курьеры ('courier_dtype = 'Авто'):"])
ws.append([
    "- Коэффициент: -0.01199",
    "- Каждая дополнительная смена снижает долю просроченных доставок на 1.2%",
    "- P-значение: 1.29577E-27 (Значимость эффекта очень высока)",
    "- R²: 0.006896 (Модель объясняет около 0.69% изменчивости)"
])

ws.append(["Электровелосипедисты ('courier_dtype = 'Электровело'):"])
ws.append([
    "- Коэффициент: -0.01362",
    "- Каждая дополнительная смена снижает долю просроченных доставок на 1.36%",
    "- P-значение: 1.07492E-07 (Значимость эффекта очень высока)",
    "- R²: 0.007945 (Модель объясняет около 0.79% изменчивости)"
])

ws.append(["Пешие курьеры ('courier_dtype = 'Пеший'):"])
ws.append([
    "- Коэффициент: -0.005186",
    "- Каждая дополнительная смена снижает долю просроченных доставок на 0.52%",
    "- P-значение: 0.54916 (Эффект статистически незначим)",
    "- R²: 0.000425 (Модель объясняет около 0.04% изменчивости)"
])

ws.append(["Велосипедисты ('courier_dtype = 'Вело'):"])
ws.append([
    "- Коэффициент: -0.02912",
    "- Каждая дополнительная смена снижает долю просроченных доставок на 2.91%",
    "- P-значение: 0.000646 (Значимость эффекта высока)",
    "- R²: 0.011508 (Модель объясняет около 1.15% изменчивости)"
])

# Автоподбор ширины столбцов
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# Добавляем графики под таблицей
start_row = len(results_df) + 5 + 10  # Увеличили отступ для добавленных выводов

for idx, (dtype, buf) in enumerate(graph_buffers):
    img = XLImage(buf)
    ws.column_dimensions['A'].width = 20
    ws.row_dimensions[start_row + idx * 20].height = 200  # Высота ячейки под графиком

    ws.cell(row=start_row + idx * 20, column=1, value=f'График: {dtype}')
    ws.merge_cells(start_row=start_row + idx * 20, start_column=1, end_row=start_row + idx * 20, end_column=5)

    ws.add_image(img, f'A{start_row + idx * 20 + 1}')

# Сохраняем файл
wb.save(output_path)

print(f"Результаты успешно сохранены в {output_path}")